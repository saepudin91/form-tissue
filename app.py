import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

# =============================
# 🔑 Koneksi Google Sheets
# =============================
scope = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]
creds = Credentials.from_service_account_info(
    st.secrets["gcp_service_account"],
    scopes=scope
)
client = gspread.authorize(creds)
sheet = client.open("Log Tissue").sheet1

# =============================
# ✅ Tambahkan Header Otomatis
# =============================
expected_header = ["Jenis", "Tanggal", "Hari", "Shift", "Pengeluaran", "Pemasukan"]
existing_values = sheet.get_all_values()

if not existing_values or existing_values[0] != expected_header:
    sheet.clear()
    sheet.append_row(expected_header)

# =============================
# 📝 Konfigurasi Aplikasi
# =============================
st.set_page_config(page_title="Form Tissue", layout="wide")
st.title("📝 Form Tissue")

# =============================
# 💾 Inisialisasi Session State
# =============================
if "data" not in st.session_state:
    st.session_state.data = []

# =============================
# 📥 Form Input
# =============================
with st.form("form_input_shift"):
    st.subheader("Input Data Tissue")
    col1, col2 = st.columns(2)

    with col1:
        jenis = st.selectbox("Jenis Tissue", [
            "Pengeluaran Tissue Roll",
            "Pengeluaran Hand Towel",
            "Pemasukan Tissue Roll",
            "Pemasukan Hand Towel"
        ])
        tanggal = st.date_input("Tanggal", value=datetime.today())
        shift = st.selectbox("Pilih Shift", ["Shift 1", "Shift 2"])

    with col2:
        jumlah_input = st.text_input("Jumlah (contoh: 1 pcs / 2 dus / 3 roll)")

    submitted = st.form_submit_button("➕ Tambahkan")

    if submitted:
        jumlah_angka = ''.join(filter(str.isdigit, jumlah_input))
        jumlah_angka = int(jumlah_angka) if jumlah_angka else 0

        pengeluaran = jumlah_angka if "Pengeluaran" in jenis else 0
        pemasukan = jumlah_angka if "Pemasukan" in jenis else 0
        jenis_bersih = jenis.replace("Pengeluaran ", "").replace("Pemasukan ", "")

        data_baris = [jenis_bersih, tanggal.strftime("%Y-%m-%d"), tanggal.strftime("%A"), shift, pengeluaran, pemasukan]
        st.session_state.data.append(data_baris)

        try:
            sheet.append_row(data_baris)
            st.success(f"✅ Data {shift} berhasil ditambahkan & disimpan ke Google Sheets!")
        except Exception as e:
            st.error(f"❌ Gagal simpan ke Google Sheets: {e}")

# =============================
# 📊 Tampilkan Data Sheet
# =============================
try:
    df = pd.DataFrame(sheet.get_all_records())
    if df.empty:
        st.info("Belum ada data pada Google Sheets.")
        st.stop()

    st.write("📋 Data Tissue Masuk & Keluar:")
    st.dataframe(df, use_container_width=True)

    df["Tanggal"] = pd.to_datetime(df["Tanggal"])

    # =============================
    # Rekap Harian
    # =============================
    pengeluaran_summary = df[df["Pengeluaran"] > 0].groupby("Jenis")["Pengeluaran"].sum().reset_index()
    pengeluaran_summary.rename(columns={"Pengeluaran": "Total Pengeluaran"}, inplace=True)

    pemasukan_summary = df[df["Pemasukan"] > 0].groupby("Jenis")["Pemasukan"].sum().reset_index()
    pemasukan_summary.rename(columns={"Pemasukan": "Total Pemasukan"}, inplace=True)

    # =============================
    # Sisa Stok dari data asli
    # =============================
    stok_df = pd.merge(pemasukan_summary, pengeluaran_summary, on="Jenis", how="outer").fillna(0)
    stok_df["Sisa Stok"] = stok_df["Total Pemasukan"] - stok_df["Total Pengeluaran"]

    st.markdown("### 📦 Sisa Stok Tissue")
    st.dataframe(stok_df)

    for _, row in stok_df.iterrows():
        if row["Sisa Stok"] <= 15:
            st.warning(f"⚠️ Stok {row['Jenis']} tersisa {int(row['Sisa Stok'])}. Segera lakukan pemesanan!")

    # =============================
    # 📥 Export ke Excel
    # =============================
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet utama
        df.to_excel(writer, sheet_name="Log Tissue", index=False, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets["Log Tissue"]

        # Judul utama
        worksheet.merge_cells("A1:F1")
        cell = worksheet["A1"]
        cell.value = "📋 Data Tissue Masuk & Keluar"
        cell.font = Font(bold=True, size=20)
        cell.alignment = Alignment(horizontal="center")

        # Lebar kolom
        for col in range(1, 7):
            worksheet.column_dimensions[get_column_letter(col)].width = 18

        # Baris awal rekap
        start_rekap_row = len(df) + 6

        # Rekap Pengeluaran
        worksheet.cell(row=start_rekap_row, column=1, value="🔻 Rekap Pengeluaran Harian").font = Font(bold=True, size=14)
        for i, row in enumerate(pengeluaran_summary.values.tolist(), start_rekap_row + 1):
            worksheet.cell(row=i, column=1, value=row[0])
            worksheet.cell(row=i, column=2, value=row[1])

        # Rekap Pemasukan
        worksheet.cell(row=start_rekap_row, column=5, value="🔺 Rekap Pemasukan Harian").font = Font(bold=True, size=14)
        for i, row in enumerate(pemasukan_summary.values.tolist(), start_rekap_row + 1):
            worksheet.cell(row=i, column=5, value=row[0])
            worksheet.cell(row=i, column=6, value=row[1])

        # Sisa Stok (dari data asli)
        stok_row_start = start_rekap_row + max(len(pengeluaran_summary), len(pemasukan_summary)) + 3
        worksheet.cell(row=stok_row_start, column=1, value="📦 Sisa Stok").font = Font(bold=True, size=14)
        for i, r in enumerate(stok_df[["Jenis", "Sisa Stok"]].values.tolist(), stok_row_start + 1):
            worksheet.cell(row=i, column=1, value=r[0])
            worksheet.cell(row=i, column=2, value=int(r[1]))

    buffer.seek(0)
    st.download_button(
        "📥 Download",
        buffer.getvalue(),
        file_name="log_tissue_dan_rekap.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # =============================
    # 📈 Rekap di Streamlit
    # =============================
    st.markdown("---")
    st.subheader("📈 Rekap Harian")
    if not pengeluaran_summary.empty:
        st.write("### 🔻 Total Pengeluaran:")
        st.dataframe(pengeluaran_summary)
    if not pemasukan_summary.empty:
        st.write("### 🔺 Total Pemasukan:")
        st.dataframe(pemasukan_summary)

except Exception as e:
    st.error(f"Gagal ambil data dari Google Sheets: {e}")
